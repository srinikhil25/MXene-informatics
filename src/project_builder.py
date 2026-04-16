# -*- coding: utf-8 -*-
"""
Project Builder
================
Scans a user's raw data folder, classifies files by technique, resolves
sample identities, parses data files, and assembles a Project object.

Usage:
    from src.project_builder import build_project
    project = build_project(Path("D:/data_raw/dhivya_data"))
"""

from __future__ import annotations

import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from src.models import (
    FileEntry, FileManifest, TechniqueData, Sample, Project,
)
from src.sample_resolver import resolve_sample_hint, SampleRegistry

logger = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────────────────────────────
# File technique classification (lightweight, no heavy imports)
# ──────────────────────────────────────────────────────────────────────────────

# Extension -> technique mapping (definitive matches)
_EXT_TECHNIQUE: dict[str, str] = {
    ".xrdml": "XRD",
    ".asc":   "XRD",
    ".emsa":  "EDS",
    ".spx":   "EDS",
    ".spe":   "XPS",
}

# Extension -> might be multiple techniques (need content sniffing)
_EXT_AMBIGUOUS: set[str] = {".txt", ".csv", ".xlsx", ".xls", ".spc"}

# Image extensions
_IMAGE_EXTENSIONS: set[str] = {
    ".tif", ".tiff", ".jpg", ".jpeg", ".bmp", ".png",
}

# Skip these entirely
_SKIP_EXTENSIONS: set[str] = {
    ".db", ".ini", ".sys", ".dll", ".exe", ".docx", ".pdf",
}

# Technique keywords in directory names
_DIR_TECHNIQUE_MAP: dict[str, str] = {
    "xrd": "XRD", "xps": "XPS", "sem": "SEM", "tem": "TEM",
    "stem": "STEM", "eds": "EDS", "edx": "EDS",
    "raman": "Raman", "ftir": "FTIR",
    "uv drs": "UV-DRS", "uv_drs": "UV-DRS", "uvdrs": "UV-DRS",
    "uv-drs": "UV-DRS", "uv-vis": "UV-Vis",
    "hall": "Hall", "hall measurement": "Hall",
    "thermoelectric": "Thermoelectric",
    "thermoelectric properties": "Thermoelectric",
    "pl": "PL", "photoluminescence": "PL",
    "hrtem": "TEM", "saed": "TEM", "haadf": "STEM",
    "images": "TEM",  # TEM/Images/ subfolder
}


def _technique_from_extension(filepath: Path) -> str:
    """Classify file by extension alone."""
    ext = filepath.suffix.lower()
    if ext in _EXT_TECHNIQUE:
        return _EXT_TECHNIQUE[ext]
    if ext in _IMAGE_EXTENSIONS:
        return "image"
    if ext in _SKIP_EXTENSIONS:
        return "skip"
    if ext in _EXT_AMBIGUOUS:
        return "ambiguous"
    return "unknown"


def _technique_from_directory(filepath: Path, root: Path) -> str:
    """Infer technique from parent directory names."""
    try:
        rel = filepath.relative_to(root)
    except ValueError:
        return ""
    parts = list(rel.parts[:-1])
    for part in parts:
        low = part.lower().strip()
        if low in _DIR_TECHNIQUE_MAP:
            return _DIR_TECHNIQUE_MAP[low]
    return ""


def _sniff_technique(filepath: Path) -> str:
    """Read first few lines to determine technique for ambiguous files."""
    ext = filepath.suffix.lower()
    try:
        with open(filepath, "r", encoding="utf-8", errors="replace") as f:
            head = f.read(500)
    except OSError:
        return "unknown"

    head_lower = head.lower()

    # UV-DRS: "Wavelength nm.","R%"
    if '"wavelength' in head_lower and 'r%' in head_lower:
        return "UV-DRS"

    # XPS CSV: line 1 is "3", line 3 is element like "Cu2p3", "Su1s"
    if ext == ".csv":
        lines = head.strip().split("\n")
        if len(lines) >= 3:
            l3 = lines[2].strip()
            # XPS region names: "Cu2p3", "Se3d5", "Su1s", "Bi4f7", "C1s1"
            import re
            if re.match(r"^[A-Z][a-z]?\d[a-z]\d?$", l3):
                return "XPS"

    # XRD Rigaku .txt: "2-theta" in header
    if "2-theta" in head_lower or "2theta" in head_lower:
        return "XRD"

    # STEM metadata .txt: "$CM_FORMAT"
    if "$cm_format" in head_lower:
        return "STEM"

    # Hall .xls — can't sniff easily (binary), rely on directory
    # Thermoelectric .xlsx — same
    return "unknown"


def _classify_file(filepath: Path, root: Path) -> FileEntry:
    """Classify a single file into a FileEntry."""
    ext = filepath.suffix.lower()
    technique = _technique_from_extension(filepath)

    if technique == "ambiguous":
        # Try directory first, then content sniffing
        dir_tech = _technique_from_directory(filepath, root)
        if dir_tech:
            technique = dir_tech
        else:
            technique = _sniff_technique(filepath)
            if technique == "unknown":
                technique = _technique_from_directory(filepath, root)

    elif technique == "image":
        # Refine image technique from directory context
        dir_tech = _technique_from_directory(filepath, root)
        if dir_tech in ("TEM", "SEM", "STEM"):
            technique = dir_tech
        # Keep as "image" if no directory context

    elif technique == "unknown":
        technique = _technique_from_directory(filepath, root) or "unknown"

    # Determine if parseable
    parseable = technique in ("XRD", "XPS", "UV-DRS", "EDS", "Hall", "Thermoelectric", "Raman")

    return FileEntry(
        path=filepath,
        technique=technique,
        file_type=ext.lstrip(".") or "unknown",
        parseable=parseable,
        metadata={},
    )


# ──────────────────────────────────────────────────────────────────────────────
# Parsing dispatcher
# ──────────────────────────────────────────────────────────────────────────────

def _parse_file(entry: FileEntry) -> Optional[dict[str, Any]]:
    """
    Parse a single file using the appropriate parser.
    Returns the parser output dict, or None on failure.
    """
    ext = entry.file_type.lower()
    path_str = str(entry.path)

    try:
        if entry.technique == "XRD":
            if ext == "xrdml":
                from src.etl.panalytical_xrd_parser import parse_panalytical_xrdml
                return parse_panalytical_xrdml(path_str)
            elif ext == "asc":
                from src.etl.asc_xrd_parser import parse_asc_xrd
                return parse_asc_xrd(path_str)
            elif ext == "txt":
                from src.etl.xrd_parser import parse_rigaku_txt
                return parse_rigaku_txt(path_str)

        elif entry.technique == "XPS":
            if ext == "csv":
                from src.etl.xps_csv_parser import parse_xps_csv
                return parse_xps_csv(path_str)
            # .spe not yet supported

        elif entry.technique == "EDS":
            if ext == "emsa":
                from src.etl.eds_parser import parse_emsa
                return parse_emsa(path_str)
            elif ext == "spx":
                from src.etl.bruker_edx_parser import parse_bruker_spx
                return parse_bruker_spx(path_str)

        elif entry.technique == "UV-DRS":
            if ext == "txt":
                return _parse_uvdrs_txt(path_str)

        elif entry.technique == "Hall":
            if ext == "xls":
                return _parse_hall_xls(path_str)

        elif entry.technique == "Thermoelectric":
            if ext == "xlsx":
                return _parse_thermoelectric_xlsx(path_str)

    except Exception as e:
        logger.warning("Failed to parse %s: %s", entry.path, e)
        return None

    return None


# ──────────────────────────────────────────────────────────────────────────────
# New lightweight parsers (inline for now, will be refactored to etl/)
# ──────────────────────────────────────────────────────────────────────────────

def _parse_uvdrs_txt(filepath: str) -> dict:
    """Parse UV-DRS .txt file (Shimadzu format: 'Sample - RawData' header)."""
    import numpy as np
    wavelengths, reflectance = [], []
    sample_name = ""

    with open(filepath, "r", encoding="utf-8", errors="replace") as f:
        for i, line in enumerate(f):
            line = line.strip().strip('"')
            if i == 0:
                # First line: "CS - RawData"
                sample_name = line.replace(" - RawData", "").strip()
                continue
            if i == 1:
                # Header: "Wavelength nm.","R%"
                continue
            # Data lines: 200.00,9.680
            parts = line.split(",")
            if len(parts) == 2:
                try:
                    wavelengths.append(float(parts[0]))
                    reflectance.append(float(parts[1]))
                except ValueError:
                    continue

    return {
        "wavelength_nm": wavelengths,
        "reflectance_pct": reflectance,
        "sample_name": sample_name,
        "source_file": filepath,
        "metadata": {
            "technique": "UV-DRS",
            "measurement_type": "diffuse_reflectance",
            "wavelength_range": f"{min(wavelengths):.0f}-{max(wavelengths):.0f} nm" if wavelengths else "",
            "n_points": len(wavelengths),
        },
    }


def _parse_hall_xls(filepath: str) -> dict:
    """Parse Hall measurement .xls file (headers at row 41, data at row 42)."""
    try:
        import xlrd
    except ImportError:
        logger.warning("xlrd not installed, cannot parse Hall .xls")
        return {}

    wb = xlrd.open_workbook(filepath)
    sh = wb.sheet_by_index(0)

    # Find the header row (contains "Temperature")
    header_row = None
    for r in range(sh.nrows):
        for c in range(min(sh.ncols, 10)):
            val = sh.cell_value(r, c)
            if isinstance(val, str) and "temperature" in val.lower():
                header_row = r
                break
        if header_row is not None:
            break

    if header_row is None:
        return {}

    # Read headers
    headers = [sh.cell_value(header_row, c) for c in range(sh.ncols)]

    # Read data rows
    data = {}
    for c, h in enumerate(headers):
        if not h or not isinstance(h, str):
            continue
        col_data = []
        for r in range(header_row + 1, sh.nrows):
            val = sh.cell_value(r, c)
            if isinstance(val, (int, float)):
                col_data.append(val)
        if col_data:
            data[h] = col_data

    # Extract key columns with clean names
    result = {"source_file": filepath, "raw_columns": data, "metadata": {}}

    key_map = {
        "temperature": "temperature_C",
        "resistivity": "resistivity_ohm_cm",
        "conductivity": "conductivity_1_ohm_cm",
        "ccc bulk": "carrier_concentration_cm3",
        "mobility": "mobility_cm2_Vs",
        "avg. hall coefficient": "hall_coefficient_cm3_C",
        "sheet resistance": "sheet_resistance_ohm",
    }
    for header, values in data.items():
        h_low = header.lower()
        for pattern, clean_name in key_map.items():
            if pattern in h_low:
                result[clean_name] = values
                break

    # Determine carrier type from Hall coefficient sign
    if "hall_coefficient_cm3_C" in result:
        hc = result["hall_coefficient_cm3_C"]
        if hc:
            result["metadata"]["carrier_type"] = "p-type" if hc[0] > 0 else "n-type"

    result["metadata"]["technique"] = "Hall"
    return result


def _parse_thermoelectric_xlsx(filepath: str) -> dict:
    """
    Parse thermoelectric .xlsx with multiple sheets (one per sample).
    Returns a dict keyed by sheet name -> data.
    """
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl not installed, cannot parse .xlsx")
        return {}

    wb = openpyxl.load_workbook(filepath, data_only=True)
    all_sheets = {}

    for sheet_name in wb.sheetnames:
        sh = wb[sheet_name]

        # Find header row (contains "Temperature")
        header_row = None
        for r in range(1, min(sh.max_row + 1, 10)):
            for c in range(1, min(sh.max_column + 1, 15)):
                val = sh.cell(r, c).value
                if isinstance(val, str) and "temperature" in val.lower():
                    header_row = r
                    break
            if header_row is not None:
                break

        if header_row is None:
            continue

        # Read headers
        headers = []
        for c in range(1, sh.max_column + 1):
            val = sh.cell(header_row, c).value
            headers.append(str(val).strip() if val else "")

        # Read data (including unnamed columns as _col_N)
        data = {}
        for ci, h in enumerate(headers):
            key = h if h else f"_col_{ci}"
            data[key] = []
        for r in range(header_row + 1, sh.max_row + 1):
            for ci, h in enumerate(headers):
                key = h if h else f"_col_{ci}"
                val = sh.cell(r, ci + 1).value
                if isinstance(val, (int, float)):
                    data[key].append(val)

        # Clean column names
        result = {"source_file": filepath, "sheet_name": sheet_name}
        col_map = {
            "temperature": "temperature_K",
            "resistivity": "resistivity_ohm_m",
            "thermal conductivity": "thermal_conductivity_W_mK",
            "seebeck": "seebeck_uV_K",
            "powerfactor": "power_factor",
            "zt": "zT",
        }
        for col_header, values in data.items():
            if col_header.startswith("_col_"):
                continue  # skip unnamed columns (used only for sanity checks)
            h_low = col_header.lower().strip()
            matched = False
            for pattern, clean_name in col_map.items():
                if pattern in h_low:
                    result[clean_name] = values
                    matched = True
                    break
            if not matched and values:
                result[col_header] = values

        # Sanity check: zT should be 0-10 for real materials.
        # Some spreadsheets have intermediate calculations in the "zT" column;
        # the actual zT may be in an adjacent unlabeled column.
        if "zT" in result and result["zT"]:
            import statistics
            median_zt = statistics.median(result["zT"])
            if median_zt > 10:
                n_expected = len(result["zT"])
                # Look through unnamed columns for one with reasonable zT values
                for key, values in data.items():
                    if not key.startswith("_col_"):
                        continue
                    if not values or len(values) != n_expected:
                        continue
                    candidate_median = statistics.median(values)
                    if 0 < candidate_median < 10:
                        logger.info(
                            f"Sheet '{sheet_name}': zT column had unphysical values "
                            f"(median={median_zt:.1f}), using adjacent unlabeled "
                            f"column (median={candidate_median:.4f}) instead."
                        )
                        result["zT"] = values
                        break

        result["metadata"] = {
            "technique": "Thermoelectric",
            "n_temperatures": len(result.get("temperature_K", [])),
        }
        all_sheets[sheet_name] = result

    return {"sheets": all_sheets, "source_file": filepath}


# ──────────────────────────────────────────────────────────────────────────────
# Main builder
# ──────────────────────────────────────────────────────────────────────────────

def build_project(
    root: Path,
    project_name: str = "",
    progress_callback=None,
) -> Project:
    """
    Scan a data folder and build a Project object.

    Parameters
    ----------
    root : Path
        Root directory to scan.
    project_name : str
        Project name. Defaults to the folder name.
    progress_callback : callable, optional
        Called with (message: str, fraction: float) for progress updates.
        Useful for Streamlit spinners/progress bars.

    Returns
    -------
    Project
        Fully assembled project with samples, techniques, and parsed data.
    """
    root = Path(root).resolve()
    if not project_name:
        project_name = root.name

    def _progress(msg: str, frac: float = 0.0):
        if progress_callback:
            progress_callback(msg, frac)

    # ── Phase 1: Scan and classify all files ─────────────────────────
    _progress("Scanning files...", 0.0)
    entries: list[FileEntry] = []
    all_files = []
    for dirpath, dirs, files in os.walk(root):
        # Skip hidden directories
        dirs[:] = [d for d in dirs if not d.startswith(".")]
        for fn in files:
            if fn.startswith("."):
                continue
            all_files.append(Path(dirpath) / fn)

    total = len(all_files)
    for i, fp in enumerate(all_files):
        entry = _classify_file(fp, root)
        if entry.technique != "skip":
            entries.append(entry)
        if i % 20 == 0:
            _progress(f"Classifying files... ({i}/{total})", 0.1 + 0.2 * i / max(total, 1))

    manifest = FileManifest(
        root=root,
        entries=entries,
        scan_time=datetime.now(),
        total_files=total,
        skipped=total - len(entries),
    )

    # ── Phase 2: Resolve sample identities ───────────────────────────
    _progress("Resolving sample identities...", 0.3)
    registry = SampleRegistry()
    for entry in entries:
        hint = resolve_sample_hint(entry.path, root)
        entry.sample_hint = hint
        registry.add_hint(hint)

    registry.resolve()

    # Apply canonical IDs back to entries
    for entry in entries:
        entry.sample_hint = registry.get_canonical(entry.sample_hint)

    # ── Phase 3: Build sample structure ──────────────────────────────
    _progress("Building sample structure...", 0.4)
    samples: dict[str, Sample] = {}
    unassigned: list[FileEntry] = []

    for entry in entries:
        sid = entry.sample_hint
        if not sid:
            unassigned.append(entry)
            continue

        if sid not in samples:
            samples[sid] = Sample(
                sample_id=sid,
                aliases=registry.get_aliases(sid),
            )

        sample = samples[sid]
        technique = entry.technique

        # Normalize image technique to the parent technique
        if technique == "image":
            technique = "TEM"  # default for unclassified images

        if technique not in sample.techniques:
            sample.techniques[technique] = TechniqueData(technique=technique)

        sample.techniques[technique].files.append(entry.path)

    # ── Phase 4: Parse data files ────────────────────────────────────
    _progress("Parsing data files...", 0.5)
    parseable = [e for e in entries if e.parseable and e.sample_hint]
    for i, entry in enumerate(parseable):
        sid = entry.sample_hint
        if sid not in samples:
            continue

        parsed = _parse_file(entry)
        if not parsed:
            continue

        technique = entry.technique
        td = samples[sid].techniques.get(technique)
        if td is None:
            continue

        # Merge parsed data into TechniqueData
        # For XPS: multiple files (survey, Cu 2p, Se 3d) -> merge by region
        if technique == "XPS" and "region" in parsed:
            region = parsed["region"]
            if "regions" not in td.parsed:
                td.parsed["regions"] = {}
            td.parsed["regions"][region] = parsed
        elif technique == "XRD":
            # Each XRD file is a separate pattern
            sample_key = entry.path.stem
            if "patterns" not in td.parsed:
                td.parsed["patterns"] = {}
            td.parsed["patterns"][sample_key] = parsed
        elif technique == "EDS":
            if "spectra" not in td.parsed:
                td.parsed["spectra"] = []
            td.parsed["spectra"].append(parsed)
        elif technique == "UV-DRS":
            # One file per sample for UV-DRS
            td.parsed.update(parsed)
        elif technique == "Hall":
            td.parsed.update(parsed)
        else:
            td.parsed.update(parsed)

        _progress(
            f"Parsing {entry.path.name}...",
            0.5 + 0.4 * (i + 1) / max(len(parseable), 1),
        )

    # ── Phase 5: Handle multi-sample files (thermoelectric xlsx) ─────
    _progress("Processing shared data files...", 0.9)
    for entry in unassigned[:]:
        if entry.technique == "Thermoelectric" and entry.file_type == "xlsx":
            parsed = _parse_file(entry)
            if parsed and "sheets" in parsed:
                for sheet_name, sheet_data in parsed["sheets"].items():
                    # Resolve sheet name to canonical sample ID
                    registry.add_hint(sheet_name)
                    registry.resolve()
                    sid = registry.get_canonical(sheet_name)
                    if not sid:
                        # Try common patterns: "CSCBI-1" -> "CS-1"
                        from src.sample_resolver import _match_irregular, SampleRegistry as SR
                        matched = _match_irregular(sheet_name)
                        sid = matched if matched else sheet_name.upper()

                    if sid not in samples:
                        samples[sid] = Sample(sample_id=sid, aliases=[sheet_name])

                    sample = samples[sid]
                    if "Thermoelectric" not in sample.techniques:
                        sample.techniques["Thermoelectric"] = TechniqueData(
                            technique="Thermoelectric"
                        )
                    sample.techniques["Thermoelectric"].files.append(entry.path)
                    sample.techniques["Thermoelectric"].parsed.update(sheet_data)

                unassigned.remove(entry)

    _progress("Done!", 1.0)

    return Project(
        name=project_name,
        root_path=root,
        samples=dict(sorted(samples.items())),
        manifest=manifest,
        unassigned=unassigned,
        created_at=datetime.now(),
    )
